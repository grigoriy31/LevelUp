import openpyxl

def spb_part_1():
    print('Обработка файла:spb_part_1')
    data = openpyxl.load_workbook('spb_part_1.xlsx')
    sheet = data.active
    row = 3
    result = sim_card()
    while True:
        name = sheet.cell(row=row, column=1).value
        if type(name) == str:
            b = result[name]
            b['gsm'] = int(sheet.cell(row=row, column=7).value)
            b['aks'] = int(sheet.cell(row=row, column=11).value)
            b['serv'] = int(sheet.cell(row=row, column=3).value)
            result[name] = b
            row += 1
        elif name == None:
            data.close()
            return result





def spb_part_2():
    print('Обработка файла:spb_part_2')
    data = openpyxl.load_workbook('spb_part_2.xlsx')
    sheet = data.active
    row = 3
    result = spb_part_1()
    while True:
        name = sheet.cell(row=row, column=1).value
        if name == None:
            data.close()
            return result
        else:
            b = result[name]
            b['gsm'] = int(sheet.cell(row=row, column=7).value)
            b['aks'] = int(sheet.cell(row=row, column=11).value)
            b['serv'] = int(sheet.cell(row=row, column=3).value)
            result[name] = b
            row += 1

def sim_card():
    print('Обработка файла:sim_cards')
    data = openpyxl.load_workbook('sim_cards.xlsx')
    sheet = data.active
    row = 3
    result = {}
    sim_smart = ['SIM YOTA смартфон МОНО 0р.', 'SIM Yota смартфон Интернет Магазин 0 руб']
    sim_data = ['SIM Yota модем МОНО 0 руб.', 'SIM Yota модем Интернет Магазин 0 руб', 'SIM Yota комплект к оборудованию']
    while True:
        name_m = sheet.cell(row=row, column=3).value
        if name_m in result:
            if sheet.cell(row=row, column=4).value in sim_smart:
                a = int(sheet.cell(row=row, column=8).value)
                b = result[name_m]
                b['sum_sim_smart'] += a
                result[name_m] = b
                row += 1
            elif sheet.cell(row=row, column=4).value in sim_data:
                a = int(sheet.cell(row=row, column=8).value)
                b = result[name_m]
                b['sum_sim_data'] += a
                result[name_m] = b
                row += 1
            else:
                row += 1
        elif sheet.cell(row=row, column=4).value == None:
            data.close()
            return result
        else:
            result[name_m] = {'sum_sim_smart' : 0, 'sum_sim_data' : 0}
            if sheet.cell(row=row, column=4).value in sim_smart:
                a = int(sheet.cell(row=row, column=8).value)
                b = result[name_m]
                b['sum_sim_smart'] += a
                result[name_m] = b
                row += 1
            elif sheet.cell(row=row, column=4).value in sim_data:
                a = int(sheet.cell(row=row, column=8).value)
                b = result[name_m]
                b['sum_sim_data'] += a
                result[name_m] = b
                row += 1
            else:
                row += 1
