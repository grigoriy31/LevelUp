import openpyxl
from imp_statistic import sim_card
resultat_sim = sim_card()

def plans():
    stat = openpyxl.load_workbook('Statistic_Spb_2022.xlsx')
    for sim_name_n in resultat_sim.keys():
        sheet = stat[sim_name_n]
        i = 2
        while True:
            data_null = sheet.cell(row=3, column=i).value
            if data_null == None:
                for row_add in range(3, 8):
                    column_out = i - 4
                    column_fact = i - 2
                    column_new = i
                    if sheet.cell(row=row_add, column=column_out).value == None or sheet.cell(row=row_add, column=column_fact).value == None:
                        row_add += 1
                    else:
                        plans_out, fact = int(sheet.cell(row=row_add, column=column_out).value), int(sheet.cell(row=row_add, column=column_fact).value)
                        if int(plans_out) > int(fact):
                            sheet.cell(row=row_add, column=column_new, value=plans_out)
                        else:
                            plans_new = fact * 1.2
                            sheet.cell(row=row_add, column=column_new, value=plans_new)
                stat.save('Statistic_Spb_2022.xlsx')
                break
            else:
                i += 1
    stat.close()






