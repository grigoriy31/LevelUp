from imp_statistic import spb_part_2
import openpyxl
result = spb_part_2()

def resultat():
    print('Занесение факта в таблицу')
    stat = openpyxl.load_workbook('Statistic_Spb_2022.xlsx')

    for name_m in result.keys():
        sheet = stat[name_m]
        column = 2
        while True:
            a = sheet.cell(row=3, column=column).value
            if a == None:
                b = result[name_m]
                sheet.cell(row=3, column=column, value=b['gsm'])
                sheet.cell(row=4, column=column, value=b['aks'])
                sheet.cell(row=5, column=column, value=b['sum_sim_smart'])
                sheet.cell(row=6, column=column, value=b['sum_sim_data'])
                sheet.cell(row=7, column=column, value=b['serv'])
                stat.save('Statistic_Spb_2022.xlsx')
                break
            else:
                column += 1
    stat.close()


